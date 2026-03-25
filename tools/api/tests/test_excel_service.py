"""ExcelService のテスト。"""

import io

import openpyxl
from services.excel_service import extract_as_text, extract_cells, fill_template


def _make_excel(data: dict[str, list[list]]) -> bytes:
    """テスト用Excelファイルを作成する。"""
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in data.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


class TestExtractCells:
    def test_extracts_simple_data(self):
        data = _make_excel({"Sheet1": [["A1", "B1"], ["A2", "B2"]]})
        result = extract_cells(data)
        assert "Sheet1" in result["sheets"]
        assert result["sheets"]["Sheet1"][0] == ["A1", "B1"]
        assert result["sheets"]["Sheet1"][1] == ["A2", "B2"]

    def test_handles_empty_cells(self):
        data = _make_excel({"Sheet1": [["A1", None, "C1"]]})
        result = extract_cells(data)
        assert result["sheets"]["Sheet1"][0] == ["A1", "", "C1"]

    def test_handles_multiple_sheets(self):
        data = _make_excel(
            {
                "Sheet1": [["data1"]],
                "Sheet2": [["data2"]],
            }
        )
        result = extract_cells(data)
        assert "Sheet1" in result["sheets"]
        assert "Sheet2" in result["sheets"]

    def test_handles_numeric_values(self):
        data = _make_excel({"Sheet1": [[42, 3.14]]})
        result = extract_cells(data)
        assert result["sheets"]["Sheet1"][0] == ["42", "3.14"]


class TestExtractAsText:
    def test_produces_readable_text(self):
        data = _make_excel({"契約書": [["派遣先", "JBS"], ["就業場所", "虎ノ門ヒルズ"]]})
        text = extract_as_text(data)
        assert "シート: 契約書" in text
        assert "JBS" in text
        assert "虎ノ門ヒルズ" in text

    def test_skips_empty_rows(self):
        data = _make_excel({"Sheet1": [[None, None], ["A", "B"]]})
        text = extract_as_text(data)
        assert "Row1" not in text
        assert "Row2" in text


class TestFillTemplate:
    def test_fills_placeholders(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "派遣先: {{company}}"
        ws["A2"] = "場所: {{location}}"
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        template_bytes = buf.getvalue()

        result_bytes = fill_template(
            template_bytes,
            {
                "company": "JBS",
                "location": "虎ノ門ヒルズ",
            },
        )

        result_wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        ws = result_wb.active
        assert ws["A1"].value == "派遣先: JBS"
        assert ws["A2"].value == "場所: 虎ノ門ヒルズ"
        result_wb.close()

    def test_preserves_cells_without_placeholders(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "通常のテキスト"
        ws["A2"] = "{{field}}"
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        result_bytes = fill_template(buf.getvalue(), {"field": "値"})
        result_wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        ws = result_wb.active
        assert ws["A1"].value == "通常のテキスト"
        assert ws["A2"].value == "値"
        result_wb.close()
