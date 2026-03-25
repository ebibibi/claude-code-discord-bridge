"""Excel 読み書きサービス。openpyxl でファイルを解析し、テンプレートに値を埋め込む。"""

from __future__ import annotations

import io
from typing import Any

import openpyxl


def extract_cells(file_bytes: bytes) -> dict[str, Any]:
    """Excelファイルの全シートから全セルの値を抽出する。

    Returns:
        {"sheets": {"Sheet1": [["A1値", "B1値"], ["A2値", "B2値"]], ...}}
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for ws in wb.worksheets:
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_cell_to_str(c) for c in row])
        sheets[ws.title] = rows
    wb.close()
    return {"sheets": sheets}


def extract_as_text(file_bytes: bytes) -> str:
    """Excelファイルの内容をAIに渡せるテキスト形式に変換する。"""
    data = extract_cells(file_bytes)
    lines: list[str] = []
    for sheet_name, rows in data["sheets"].items():
        lines.append(f"=== シート: {sheet_name} ===")
        for i, row in enumerate(rows, 1):
            non_empty = [(j + 1, v) for j, v in enumerate(row) if v]
            if non_empty:
                cells_str = ", ".join(f"Col{j}: {v}" for j, v in non_empty)
                lines.append(f"Row{i}: {cells_str}")
        lines.append("")
    return "\n".join(lines)


def fill_template(template_bytes: bytes, field_values: dict[str, str]) -> bytes:
    """テンプレートExcelに値を埋め込んで返す。

    テンプレートのセルに {{field_name}} のプレースホルダがある場合、
    field_values[field_name] の値で置き換える。
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    for field_name, value in field_values.items():
                        placeholder = "{{" + field_name + "}}"
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, value)
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)
